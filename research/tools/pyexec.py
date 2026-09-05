"""TASK-4 H10 execution sandbox: run model-written Python against one task dir.

Contract: task_dir contains input.xlsx; the model's code runs with
cwd=task_dir and must write output.xlsx there. The code text is written to
task_dir/solution.py and executed by this venv's interpreter in isolated
mode ('-I') with a stripped environment (PATH and HOME only — no API keys).

This is a BEST-EFFORT local sandbox: a network-disabling preamble stubs
socket.socket before user code runs, but a determined program could undo it.
The submission Docker container is the real security boundary; this module
only keeps honest code honest during local research runs.

Also provides the prompt builders for the exec-code strategy:
    build_exec_prompt(instruction, manifest_json, answer_position, answer_sheet)
    extract_code(response_text)
    build_repair_prompt(prev_code, stdout, stderr, traceback_text)  # ONE repair

Self-tests: uv run tools/pyexec.py
"""

from __future__ import annotations

import os
import re
import signal
import subprocess
import sys
from pathlib import Path

OUTPUT_CAP = 20_000  # max chars of stdout/stderr kept

# Prepended to solution.py before the model's code. Best-effort network
# block only; nothing else is pre-imported.
NETWORK_PREAMBLE = '''\
# --- sandbox preamble: best-effort network disable (do not edit) ---
import socket as _socket
def _network_disabled(*_args, **_kwargs):
    raise RuntimeError("network disabled")
_socket.socket = _network_disabled
_socket.create_connection = _network_disabled
_socket.socketpair = _network_disabled
_socket.getaddrinfo = _network_disabled
del _socket, _network_disabled
# --- end sandbox preamble ---
'''


def _cap(text: str) -> str:
    if text is None:
        return ""
    if len(text) > OUTPUT_CAP:
        return text[:OUTPUT_CAP] + f"\n[... truncated at {OUTPUT_CAP} chars]"
    return text


def run_code(code: str, task_dir: Path, timeout: int = 60) -> dict:
    """Write code to task_dir/solution.py (with the network preamble) and run
    it in a subprocess: this venv's python, '-I' isolated mode, cwd=task_dir,
    env stripped to PATH/HOME. Kills the whole process group on timeout.

    Returns {ok, stdout, stderr, returncode, output_exists} where ok means
    exit code 0 AND task_dir/output.xlsx exists.
    """
    task_dir = Path(task_dir)
    solution = task_dir / "solution.py"
    solution.write_text(NETWORK_PREAMBLE + "\n" + code)

    env = {
        "PATH": os.environ.get("PATH", "/usr/bin:/bin"),
        "HOME": os.environ.get("HOME", str(task_dir)),
    }
    proc = subprocess.Popen(
        [sys.executable, "-I", str(solution)],
        cwd=str(task_dir),
        env=env,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
        start_new_session=True,  # own process group, so timeout kill is total
    )
    timed_out = False
    try:
        stdout, stderr = proc.communicate(timeout=timeout)
    except subprocess.TimeoutExpired:
        timed_out = True
        try:
            os.killpg(proc.pid, signal.SIGKILL)
        except (ProcessLookupError, PermissionError):
            proc.kill()
        stdout, stderr = proc.communicate()
        stderr = (stderr or "") + f"\n[killed: timeout after {timeout}s]"

    output_exists = (task_dir / "output.xlsx").exists()
    returncode = proc.returncode
    ok = (not timed_out) and returncode == 0 and output_exists
    return {
        "ok": ok,
        "stdout": _cap(stdout),
        "stderr": _cap(stderr),
        "returncode": returncode,
        "output_exists": output_exists,
    }


# ------------------------------------------------------------------ prompts

EXEC_PROMPT_TEMPLATE = """\
You are an expert spreadsheet engineer. Solve the task below by writing a
single Python script.

## Task instruction
{instruction}

## Workbook manifest (structure of input.xlsx)
{manifest_json}

## Where the answer goes
Sheet: {answer_sheet}
Position: {answer_position}

## Requirements
- Reply with exactly ONE fenced code block: ```python ... ```
- The script runs with its working directory containing input.xlsx.
- Load the workbook with openpyxl (`openpyxl.load_workbook("input.xlsx")`).
- Perform the required transformation, preserving all unrelated content.
- Write the result to output.xlsx in the working directory
  (`wb.save("output.xlsx")`).
- Use only the Python standard library and openpyxl. No network access.
- Do not print the answer instead of writing it; output.xlsx is the answer.
"""


def build_exec_prompt(instruction: str, manifest_json: str,
                      answer_position: str, answer_sheet: str) -> str:
    """Prompt asking the model for one python fence that transforms
    input.xlsx into output.xlsx."""
    return EXEC_PROMPT_TEMPLATE.format(
        instruction=instruction,
        manifest_json=manifest_json,
        answer_position=answer_position,
        answer_sheet=answer_sheet,
    )


REPAIR_PROMPT_TEMPLATE = """\
Your previous script failed. Fix it and reply with exactly ONE fenced
```python ... ``` code block containing the complete corrected script.
Same contract: load input.xlsx with openpyxl, write output.xlsx in the
working directory. This is the only repair attempt.

## Previous script
```python
{prev_code}
```

## stdout
{stdout}

## stderr
{stderr}

## traceback
{traceback_text}
"""


def build_repair_prompt(prev_code: str, stdout: str, stderr: str,
                        traceback_text: str) -> str:
    """ONE-repair prompt: previous code plus its failure evidence."""
    return REPAIR_PROMPT_TEMPLATE.format(
        prev_code=prev_code,
        stdout=stdout or "(empty)",
        stderr=stderr or "(empty)",
        traceback_text=traceback_text or "(none)",
    )


_FENCE_PY = re.compile(r"```python\s*\n(.*?)```", re.DOTALL | re.IGNORECASE)
_FENCE_ANY = re.compile(r"```\s*\n(.*?)```", re.DOTALL)


def extract_code(response_text: str) -> str:
    """First ```python fence; else first bare fence; else the whole text."""
    if not response_text:
        return ""
    m = _FENCE_PY.search(response_text)
    if m:
        return m.group(1).strip()
    m = _FENCE_ANY.search(response_text)
    if m:
        return m.group(1).strip()
    return response_text.strip()


# --------------------------------------------------------------- self-tests

def _selftest() -> int:  # pragma: no cover
    import shutil
    import tempfile

    import openpyxl

    failures = []

    def check(name, cond, detail=""):
        status = "PASS" if cond else "FAIL"
        print(f"[{status}] {name}" + (f" — {detail}" if detail and not cond else ""))
        if not cond:
            failures.append(name)

    def fresh_task_dir(root: Path, name: str) -> Path:
        d = root / name
        d.mkdir()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "header"
        ws["A2"] = 42
        wb.save(d / "input.xlsx")
        return d

    root = Path(tempfile.mkdtemp(prefix="pyexec_selftest_"))
    try:
        # (a) happy path: copy input.xlsx -> output.xlsx via openpyxl
        d = fresh_task_dir(root, "happy")
        res = run_code(
            "import openpyxl\n"
            "wb = openpyxl.load_workbook('input.xlsx')\n"
            "wb.save('output.xlsx')\n"
            "print('copied')\n",
            d, timeout=60,
        )
        check("happy path ok", res["ok"] is True, repr(res))
        check("happy path stdout", "copied" in res["stdout"], repr(res["stdout"]))
        check("happy path output_exists", res["output_exists"] is True)

        # (b) timeout: infinite loop is killed
        d = fresh_task_dir(root, "timeout")
        res = run_code("while True:\n    pass\n", d, timeout=3)
        check("timeout ok is False", res["ok"] is False, repr(res))
        check("timeout reported", "timeout" in res["stderr"].lower(), repr(res["stderr"]))

        # (c) crash: exception captured in stderr
        d = fresh_task_dir(root, "crash")
        res = run_code("raise ValueError('boom-message')\n", d, timeout=30)
        check("crash ok is False", res["ok"] is False, repr(res))
        check("crash returncode nonzero", res["returncode"] not in (0, None), repr(res["returncode"]))
        check("crash stderr captured", "boom-message" in res["stderr"], repr(res["stderr"]))

        # (d) network: socket use raises RuntimeError from the preamble
        d = fresh_task_dir(root, "network")
        res = run_code(
            "import socket\n"
            "socket.socket(socket.AF_INET, socket.SOCK_STREAM)\n",
            d, timeout=30,
        )
        check("network ok is False", res["ok"] is False, repr(res))
        check("network blocked", "network disabled" in res["stderr"], repr(res["stderr"]))

        # (e) prompt helpers
        code = extract_code("text before\n```python\nprint(1)\n```\nafter")
        check("extract_code python fence", code == "print(1)", repr(code))
        code = extract_code("```\nx = 2\n```")
        check("extract_code bare fence", code == "x = 2", repr(code))
        code = extract_code("just plain code")
        check("extract_code fallback", code == "just plain code", repr(code))
        prompt = build_exec_prompt("Sum column B", "{}", "C1", "Sheet1")
        check("build_exec_prompt fields",
              all(s in prompt for s in ("Sum column B", "C1", "Sheet1", "```python")))
        rp = build_repair_prompt("print(1)", "", "Trace...", "ValueError")
        check("build_repair_prompt fields",
              all(s in rp for s in ("print(1)", "Trace...", "ValueError")))
    finally:
        shutil.rmtree(root, ignore_errors=True)

    print()
    if failures:
        print(f"SELF-TEST FAILED: {len(failures)} failing check(s): {failures}")
        return 1
    print("SELF-TEST OK: all checks passed")
    return 0


if __name__ == "__main__":
    sys.exit(_selftest())
