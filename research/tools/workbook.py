"""TASK-5 workbook inspection tooling.

Read-only helpers over a single INPUT workbook path (the task's init file).
These functions are for building model context and debugging; they operate
only on the path they are given and never resolve any other file.

All return values are JSON-serializable and size-capped: strings are
truncated to MAX_STR chars, lists are capped by the constants below.

API:
    build_manifest(xlsx_path) -> dict
    read_range(path, sheet, a1range) -> list[list]
    inspect_rows(path, sheet, start, end) -> list[list]
    inspect_columns(path, sheet, columns) -> dict[str, list]
    find_text(path, query, max_hits=20) -> list[[sheet, coord, value]]
    inspect_formulas(path, sheet, a1range) -> list[[coord, formula]]
"""

from __future__ import annotations

import datetime
from pathlib import Path

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter, range_boundaries

MAX_STR = 120          # truncate any string value to this many chars
MAX_ROWS = 200         # cap on rows returned by range/row readers
MAX_COLS = 60          # cap on columns returned per row
MAX_COL_VALUES = 500   # cap on values returned per column by inspect_columns
MAX_LIST = 100         # generic cap for manifest lists (merged ranges, etc.)
MAX_REGIONS = 5        # candidate data regions kept per sheet
HEADER_TEXT_FRACTION = 0.6
MANIFEST_SCAN_ROWS = 20000   # safety cap on rows scanned per sheet
MANIFEST_SCAN_COLS = 200     # safety cap on columns scanned per sheet


# ---------------------------------------------------------------- utilities

def _clean(value):
    """Make a cell value JSON-serializable and cap string length."""
    if value is None or isinstance(value, (bool, int)):
        return value
    if isinstance(value, float):
        return value if value == value and abs(value) != float("inf") else str(value)
    if isinstance(value, str):
        return value if len(value) <= MAX_STR else value[:MAX_STR] + "…"
    if isinstance(value, (datetime.datetime, datetime.date, datetime.time)):
        return value.isoformat()
    if isinstance(value, datetime.timedelta):
        return str(value)
    return str(value)[:MAX_STR]


def _load(path, data_only=True):
    return openpyxl.load_workbook(Path(path), data_only=data_only, read_only=False)


def _get_sheet(wb, sheet):
    if sheet not in wb.sheetnames:
        raise KeyError(f"sheet {sheet!r} not found; workbook has {wb.sheetnames}")
    return wb[sheet]


def _cell_type(value):
    """Coarse type tag used for header inference."""
    if value is None:
        return "empty"
    if isinstance(value, str):
        return "formula" if value.startswith("=") else "text"
    if isinstance(value, bool):
        return "bool"
    if isinstance(value, (int, float)):
        return "number"
    if isinstance(value, (datetime.datetime, datetime.date, datetime.time)):
        return "date"
    return "other"


# ------------------------------------------------------------ manifest bits

def _scan_grid(ws):
    """One pass over the used grid: per-row non-empty counts, per-row type
    profiles (first rows only), per-column non-empty and formula counts."""
    max_row = min(ws.max_row or 0, MANIFEST_SCAN_ROWS)
    max_col = min(ws.max_column or 0, MANIFEST_SCAN_COLS)
    row_nonempty = {}          # row index -> count of non-empty cells
    row_types = {}             # row index -> list of type tags (first 50 rows)
    col_nonempty = {}          # col index -> count
    col_formulas = {}          # col index -> count of '=' cells
    min_r = min_c = None
    max_r = max_c = 0
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
        for cell in row:
            v = cell.value
            if v is None or v == "":
                continue
            r, c = cell.row, cell.column
            row_nonempty[r] = row_nonempty.get(r, 0) + 1
            col_nonempty[c] = col_nonempty.get(c, 0) + 1
            if isinstance(v, str) and v.startswith("="):
                col_formulas[c] = col_formulas.get(c, 0) + 1
            if r <= 50:
                row_types.setdefault(r, []).append(_cell_type(v))
            if min_r is None or r < min_r:
                min_r = r
            if min_c is None or c < min_c:
                min_c = c
            max_r = max(max_r, r)
            max_c = max(max_c, c)
    used = None
    if min_r is not None:
        used = f"{get_column_letter(min_c)}{min_r}:{get_column_letter(max_c)}{max_r}"
    return {
        "row_nonempty": row_nonempty,
        "row_types": row_types,
        "col_nonempty": col_nonempty,
        "col_formulas": col_formulas,
        "used_range": used,
    }


def _infer_header_row(row_types):
    """First row where >=60% of non-empty cells are text AND the next
    populated row has a different type profile."""
    rows = sorted(row_types)
    for i, r in enumerate(rows):
        tags = row_types[r]
        if not tags:
            continue
        text_frac = sum(1 for t in tags if t == "text") / len(tags)
        if text_frac < HEADER_TEXT_FRACTION:
            continue
        if i + 1 < len(rows):
            nxt = row_types[rows[i + 1]]
            profile = sorted(set(tags))
            nxt_profile = sorted(set(nxt))
            if profile != nxt_profile or len(tags) != len(nxt):
                return r
        else:
            return r
    return None


def _data_regions(row_nonempty):
    """Contiguous non-empty row blocks, top MAX_REGIONS by size."""
    rows = sorted(row_nonempty)
    regions = []
    start = prev = None
    for r in rows:
        if prev is not None and r == prev + 1:
            prev = r
            continue
        if start is not None:
            regions.append({"start_row": start, "end_row": prev, "rows": prev - start + 1})
        start = prev = r
    if start is not None:
        regions.append({"start_row": start, "end_row": prev, "rows": prev - start + 1})
    regions.sort(key=lambda x: (-x["rows"], x["start_row"]))
    return regions[:MAX_REGIONS]


def build_manifest(xlsx_path) -> dict:
    """Structural manifest of one workbook: sheets, sizes, used ranges,
    named ranges, tables, merged ranges, per-column formula density,
    inferred header row, and candidate data regions."""
    path = Path(xlsx_path)
    wb = _load(path, data_only=False)
    manifest = {
        "path": str(path),
        "sheet_names": wb.sheetnames[:MAX_LIST],
        "named_ranges": {},
        "sheets": {},
    }
    try:
        for name, dn in list(wb.defined_names.items())[:MAX_LIST]:
            manifest["named_ranges"][str(name)[:MAX_STR]] = _clean(dn.value)
    except Exception:
        pass

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        scan = _scan_grid(ws)
        col_density = {}
        for c, n_formulas in sorted(scan["col_formulas"].items())[:MAX_LIST]:
            total = scan["col_nonempty"].get(c, 0)
            if total:
                col_density[get_column_letter(c)] = round(n_formulas / total, 4)
        tables = []
        try:
            for tname, tbl in list(ws.tables.items())[:MAX_LIST]:
                tables.append({"name": _clean(tname), "ref": _clean(tbl.ref)})
        except Exception:
            pass
        merged = [str(m) for m in list(ws.merged_cells.ranges)[:MAX_LIST]]
        manifest["sheets"][sheet_name] = {
            "max_row": ws.max_row,
            "max_col": ws.max_column,
            "dimensions": ws.dimensions,
            "used_range_estimate": scan["used_range"],
            "nonempty_row_count": len(scan["row_nonempty"]),
            "nonempty_col_count": len(scan["col_nonempty"]),
            "tables": tables,
            "merged_ranges": merged,
            "formula_density_by_column": col_density,
            "inferred_header_row": _infer_header_row(scan["row_types"]),
            "candidate_data_regions": _data_regions(scan["row_nonempty"]),
        }
    wb.close()
    return manifest


# ------------------------------------------------------------- cell readers

def _rows_from_bounds(ws, min_col, min_row, max_col, max_row):
    max_row = min(max_row, min_row + MAX_ROWS - 1)
    max_col = min(max_col, min_col + MAX_COLS - 1)
    out = []
    for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                            min_col=min_col, max_col=max_col):
        out.append([_clean(cell.value) for cell in row])
    return out


def read_range(path, sheet, a1range):
    """Values in an A1-style range (e.g. 'A1:D20'), capped to
    MAX_ROWS x MAX_COLS. Returns a list of rows of values."""
    wb = _load(path, data_only=True)
    ws = _get_sheet(wb, sheet)
    min_col, min_row, max_col, max_row = range_boundaries(a1range)
    out = _rows_from_bounds(ws, min_col or 1, min_row or 1,
                            max_col or ws.max_column, max_row or ws.max_row)
    wb.close()
    return out


def inspect_rows(path, sheet, start, end):
    """Values for rows start..end inclusive (1-based), capped."""
    wb = _load(path, data_only=True)
    ws = _get_sheet(wb, sheet)
    out = _rows_from_bounds(ws, 1, max(1, int(start)), ws.max_column or 1, int(end))
    wb.close()
    return out


def inspect_columns(path, sheet, columns):
    """Values for the given column letters, e.g. ['A', 'C'].
    Returns {letter: [values...]} capped to MAX_COL_VALUES per column."""
    wb = _load(path, data_only=True)
    ws = _get_sheet(wb, sheet)
    out = {}
    for letter in list(columns)[:MAX_COLS]:
        idx = column_index_from_string(letter)
        values = []
        for (cell,) in ws.iter_rows(min_col=idx, max_col=idx,
                                    max_row=min(ws.max_row or 1, MAX_COL_VALUES)):
            values.append(_clean(cell.value))
        out[letter] = values
    wb.close()
    return out


def find_text(path, query, max_hits=20):
    """Case-insensitive substring search over string cells in all sheets.
    Returns up to max_hits [sheet, coordinate, value] triples."""
    needle = str(query).lower()
    wb = _load(path, data_only=True)
    hits = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(max_row=min(ws.max_row or 1, MANIFEST_SCAN_ROWS),
                                max_col=min(ws.max_column or 1, MANIFEST_SCAN_COLS)):
            for cell in row:
                v = cell.value
                if isinstance(v, str) and needle in v.lower():
                    hits.append([sheet_name, cell.coordinate, _clean(v)])
                    if len(hits) >= max_hits:
                        wb.close()
                        return hits
    wb.close()
    return hits


def inspect_formulas(path, sheet, a1range):
    """[coordinate, formula] pairs for formula cells in an A1 range
    (loaded with data_only=False), capped to MAX_ROWS*MAX_COLS scan."""
    wb = _load(path, data_only=False)
    ws = _get_sheet(wb, sheet)
    min_col, min_row, max_col, max_row = range_boundaries(a1range)
    min_col, min_row = min_col or 1, min_row or 1
    max_col = min(max_col or ws.max_column, min_col + MAX_COLS - 1)
    max_row = min(max_row or ws.max_row, min_row + MAX_ROWS - 1)
    out = []
    for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                            min_col=min_col, max_col=max_col):
        for cell in row:
            v = cell.value
            if isinstance(v, str) and v.startswith("="):
                out.append([cell.coordinate, _clean(v)])
    wb.close()
    return out
