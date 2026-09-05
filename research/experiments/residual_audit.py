"""E009 residual audit: full classification of the 95 official-400 failures.

    uv run experiments/residual_audit.py ../experiments/E009-champion-400 \
        --out ../experiments/E009-residual-audit.json

Analysis-side tool: goldens inform classification via the official results
(mismatch samples, correct/cells) — never inference. Per failed task:
instruction type, selected cascade stage + health verdict, failure scale,
dominant mechanism (heuristic, documented), answer-range size, workbook dims,
and addressability tags for the五 candidate interventions.
"""

import argparse
import json
import sys
from collections import Counter
from pathlib import Path

RESEARCH = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(RESEARCH))

import openpyxl
from openpyxl.utils.cell import coordinate_to_tuple

from sb import answer_cells, load_dataset

ERRV = {"#NAME?", "#N/A", "#REF!", "#VALUE!", "#DIV/0!", "#NULL!", "#NUM!"}


def scale_of(item):
    if item.get("status") != "graded":
        return "dead"
    cells, correct = item.get("cells", 0), item.get("correct", 0)
    wrong = cells - correct
    if correct == 0:
        return "dead"
    if wrong == 1:
        return "1-cell"
    if wrong / cells < 0.10:
        return "sparse"
    if wrong / cells < 0.60:
        return "partial"
    return "broad"


def mechanism_of(item, task, traces, dims, max_tokens):
    sel = next((t for t in traces if t.get("selected")), traces[-1] if traces else {})
    mism = item.get("mismatches") or []
    if any("MergedCell" in (t.get("error") or "") for t in traces):
        return "unsupported structural transformation"
    if item.get("status") != "graded":
        return "unknown"
    acts = [m.get("actual") for m in mism]
    exps = [m.get("expected") for m in mism]
    if mism and all(str(e) == str(a) for e, a in zip(exps, acts)):
        return "wrong typing/date"
    if any(isinstance(a, str) and a.strip() in ERRV for a in acts):
        return "formula error"
    if (sel.get("output_tokens") or 0) >= max_tokens:
        return "large-output/combinatorial"
    n_answer_sheets = len({s for s, _c in answer_cells_cached(task)})
    if n_answer_sheets > 1:
        return "multi-sheet reasoning"
    empty_frac = sum(a is None for a in acts) / len(acts) if acts else 0
    if empty_frac >= 0.6:
        return "incomplete output"
    oversize = any((r or 0) > 120 or (c or 0) > 30 for r, c in dims.values())
    if oversize and any(a in (None, 0, "0") for a in acts):
        return "missing source visibility"
    numericish = 0
    for e, a in zip(exps, acts):
        try:
            float(e); float(a); numericish += 1
        except (TypeError, ValueError):
            pass
    if numericish >= max(1, len(mism) // 2):
        return "arithmetic / logic"
    if len(mism) >= 2 and all(isinstance(a, str) and isinstance(e, str) for e, a in zip(exps, acts)):
        return "wrong lookup/range"
    return "unknown"


_cells_cache = {}


def answer_cells_cached(task):
    if task["id"] not in _cells_cache:
        wb = openpyxl.load_workbook(task["init_xlsx"], read_only=True)
        _cells_cache[task["id"]] = answer_cells(task, wb)
        wb.close()
    return _cells_cache[task["id"]]


def addressability(row):
    tags = []
    if row["healthy_but_wrong"] and row["scale"] in ("1-cell", "sparse") :
        tags.append("self-consistency")
    if row["mechanism"] in ("incomplete output", "wrong typing/date", "formula error") or row["scale"] in ("1-cell", "sparse"):
        tags.append("targeted repair")
    if row["type"].startswith("Sheet") and (row["answer_cells"] > 300 or row["mechanism"] in
            ("large-output/combinatorial", "unsupported structural transformation", "missing source visibility")):
        tags.append("code execution")
    if row["mechanism"] == "missing source visibility":
        tags.append("better workbook inspection")
    if row["mechanism"] in ("arithmetic / logic", "wrong lookup/range", "multi-sheet reasoning", "unknown") and row["scale"] in ("partial", "broad", "dead"):
        tags.append("model reasoning/post-training")
    return tags or ["model reasoning/post-training"]


def main():
    p = argparse.ArgumentParser()
    p.add_argument("exp")
    p.add_argument("--out", required=True)
    args = p.parse_args()
    exp = Path(args.exp)
    results = json.loads((exp / "results.json").read_text())
    manifest = json.loads((exp / "manifest.json").read_text())
    max_tokens = manifest.get("max_tokens") or 24576
    tasks = {str(t["id"]): t for t in load_dataset()}

    rows = []
    for item in results["items"]:
        if item.get("pass"):
            continue
        tid = str(item["id"])
        task = tasks[tid]
        tf = exp / "traces" / f"{tid}.jsonl"
        traces = [json.loads(l) for l in tf.read_text().splitlines()] if tf.exists() else []
        sel = next((t for t in traces if t.get("selected")), traces[-1] if traces else {})
        wb = openpyxl.load_workbook(task["init_xlsx"], read_only=True)
        dims = {ws.title: (ws.max_row, ws.max_column) for ws in wb.worksheets}
        wb.close()
        row = {
            "id": tid,
            "type": item["type"],
            "stage": sel.get("stage", "?"),
            "healthy": bool((sel.get("health") or {}).get("healthy")),
            "scale": scale_of(item),
            "cells": item.get("cells"),
            "correct": item.get("correct"),
            "answer_cells": len(answer_cells_cached(task)),
            "dims": dims,
        }
        row["healthy_but_wrong"] = row["healthy"] and item.get("status") == "graded"
        row["mechanism"] = mechanism_of(item, task, traces, dims, max_tokens)
        row["addressable_by"] = addressability(row)
        rows.append(row)

    agg = {
        "n_failures": len(rows),
        "A_healthy_but_wrong": sum(r["healthy_but_wrong"] for r in rows),
        "B_leq5_wrong_cells": sum(1 for r in rows if r["cells"] and 0 < r["cells"] - r["correct"] <= 5),
        "C_geq90pct_cell_acc": sum(1 for r in rows if r["cells"] and r["correct"] / r["cells"] >= 0.9),
        "D_large_sheet_transform": sum(1 for r in rows if r["type"].startswith("Sheet") and r["answer_cells"] > 300),
        "by_scale": dict(Counter(r["scale"] for r in rows)),
        "by_mechanism": dict(Counter(r["mechanism"] for r in rows).most_common()),
        "by_stage": dict(Counter(r["stage"] for r in rows)),
        "by_type": dict(Counter(r["type"][:5] for r in rows)),
        "E_addressable_by": dict(Counter(t for r in rows for t in r["addressable_by"]).most_common()),
    }
    Path(args.out).write_text(json.dumps({"aggregates": agg, "rows": rows}, indent=1, default=str))
    print(json.dumps(agg, indent=2))


if __name__ == "__main__":
    main()
