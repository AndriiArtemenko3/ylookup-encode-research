"""Failure analysis over official results.json files. Stdlib only.

    uv run experiments/compare.py ../experiments/E000-x/results.json
    uv run experiments/compare.py ../experiments/E000-x/results.json ../experiments/E001-y/results.json

One argument: report on a single run (pass rates, failures by instruction type,
failed ids, graded-cell and workbook stats). Two arguments: additionally diff
them (FAIL->PASS / PASS->FAIL / PASS->PASS / FAIL->FAIL, net change).

The official results.json is the only source of truth; this script computes no
scores of its own.
"""

import argparse
import json
import sys
from collections import Counter
from pathlib import Path

RESEARCH = Path(__file__).resolve().parents[1]
DATASET_DIR = RESEARCH / "data" / "spreadsheetbench_verified_400"


def load_results(path: str) -> dict[str, dict]:
    data = json.loads(Path(path).read_text())
    return {str(i["id"]): i for i in data["items"]}, data.get("summary", {})


def load_tasks() -> dict[str, dict]:
    dataset = DATASET_DIR / "dataset.json"
    if not dataset.exists():
        return {}
    return {str(t["id"]): t for t in json.loads(dataset.read_text())}


def workbook_dims(task: dict) -> str:
    """Sheet dimensions of the init workbook, when openpyxl is available and the file exists."""
    try:
        import openpyxl
        folder = DATASET_DIR / task["spreadsheet_path"]
        init = next(folder.glob("*init*.xlsx"))
        wb = openpyxl.load_workbook(init, read_only=True)
        dims = ", ".join(f"{ws.title}:{ws.max_row}x{ws.max_column}" for ws in wb.worksheets)
        wb.close()
        return dims
    except Exception:
        return "?"


def passed(item: dict) -> bool:
    return bool(item.get("pass", False))


def report_single(items: dict[str, dict], summary: dict, tasks: dict[str, dict], dims: bool) -> None:
    print(json.dumps(summary, indent=2))
    n = len(items)
    by_status = Counter(i["status"] for i in items.values())
    print(f"\nitems {n}  status: {dict(by_status)}")
    print(f"missing: {sum(s in ('missing', 'missing_output') for s in by_status.elements())}"
          f"  errors: {by_status.get('error', 0)}")

    for level in ("Cell", "Sheet"):
        rows = [i for i in items.values() if i["type"].startswith(level)]
        if rows:
            p = sum(passed(i) for i in rows)
            print(f"{level}-level: {p}/{len(rows)} pass ({p / len(rows):.1%})")

    graded = [i for i in items.values() if "cells" in i]
    if graded:
        cell_counts = Counter(i["cells"] for i in graded)
        print(f"graded cells per task: min {min(cell_counts)} max {max(cell_counts)}"
              f" total {sum(i['cells'] for i in graded)}")

    failed = sorted((i for i in items.values() if not passed(i)), key=lambda i: i["id"])
    print(f"\nfailed ids ({len(failed)}):")
    for i in failed:
        extra = workbook_dims(tasks[i["id"]]) if dims and i["id"] in tasks else ""
        mism = i.get("mismatches") or []
        first = f" first: {mism[0]['cell']} expected {mism[0]['expected']!r} got {mism[0]['actual']!r}" if mism else ""
        print(f"  {i['id']:<8} {i['type'][:5]:<6} {i['status']:<15} "
              f"{i.get('correct', 0)}/{i.get('cells', '?')}{first} {extra}")


def report_diff(a: dict[str, dict], b: dict[str, dict], name_a: str, name_b: str) -> None:
    ids = sorted(set(a) | set(b))
    buckets: dict[str, list[str]] = {"FAIL->PASS": [], "PASS->FAIL": [], "PASS->PASS": [], "FAIL->FAIL": []}
    for task_id in ids:
        was = passed(a.get(task_id, {}))
        now = passed(b.get(task_id, {}))
        key = f"{'PASS' if was else 'FAIL'}->{'PASS' if now else 'FAIL'}"
        buckets[key].append(task_id)
    print(f"\n== diff: {name_a} -> {name_b} ==")
    for key in ("FAIL->PASS", "PASS->FAIL", "PASS->PASS", "FAIL->FAIL"):
        print(f"{key}: {len(buckets[key])}")
    net = len(buckets["FAIL->PASS"]) - len(buckets["PASS->FAIL"])
    print(f"net change: {net:+d} tasks")
    if buckets["FAIL->PASS"]:
        print(f"fixed:     {', '.join(buckets['FAIL->PASS'])}")
    if buckets["PASS->FAIL"]:
        print(f"regressed: {', '.join(buckets['PASS->FAIL'])}")


def main():
    p = argparse.ArgumentParser()
    p.add_argument("results", nargs="+", help="one or two results.json paths (from evaluate.py --out)")
    p.add_argument("--dims", action="store_true", help="include init workbook dimensions for failed tasks")
    args = p.parse_args()
    if len(args.results) > 2:
        sys.exit("give one or two results.json paths")

    tasks = load_tasks()
    items_a, summary_a = load_results(args.results[0])
    print(f"== {args.results[0]} ==")
    report_single(items_a, summary_a, tasks, args.dims)

    if len(args.results) == 2:
        items_b, summary_b = load_results(args.results[1])
        print(f"\n== {args.results[1]} ==")
        report_single(items_b, summary_b, tasks, args.dims)
        report_diff(items_a, items_b, args.results[0], args.results[1])


if __name__ == "__main__":
    main()
