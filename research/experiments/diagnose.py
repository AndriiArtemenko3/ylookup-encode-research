"""Per-failure diagnostic bundles for an experiment. Evaluation-side tooling.

    uv run experiments/diagnose.py ../experiments/E002-maxtokens-24k-dev

For every failed task it gathers the evidence needed to classify the failure:
trace error and token counts, official mismatches (expected vs actual, from
results.json — the evaluator's output, the only place golden-derived values
belong), workbook dimensions vs the 120x30 serialisation window, and the tail
of the model's reply. Prints one compact block per task, worst-first is left
to the reader. Classification itself is judgment and lives in notes.md.
"""

import json
import sys
from pathlib import Path

import openpyxl

RESEARCH = Path(__file__).resolve().parents[1]
DATASET_DIR = RESEARCH / "data" / "spreadsheetbench_verified_400"
SER_ROWS, SER_COLS = 120, 30  # sb.serialize_workbook window


def mismatch_flavor(expected, actual):
    if actual is None:
        return "model wrote nothing/empty"
    try:
        e, a = float(expected), float(actual)
        if abs(e - a) < 0.051:
            return f"NEAR MISS (diff {a - e:+.3f})"
        if e != 0 and 0.9 < a / e < 1.1:
            return "within 10%"
        return "far off"
    except (TypeError, ValueError):
        if str(expected).strip().lower() == str(actual).strip().lower():
            return "case/whitespace only"
        if type(expected) is not type(actual):
            return f"type mismatch ({type(expected).__name__} vs {type(actual).__name__})"
        return "different text"


def main():
    exp = Path(sys.argv[1])
    results = json.loads((exp / "results.json").read_text())
    tasks = {str(t["id"]): t for t in json.loads((DATASET_DIR / "dataset.json").read_text())}

    for item in results["items"]:
        if item.get("pass", False):
            continue
        tid = item["id"]
        task = tasks[tid]
        trace = json.loads((exp / "traces" / f"{tid}.jsonl").read_text())

        folder = DATASET_DIR / task["spreadsheet_path"]
        init = next(folder.glob("*init*.xlsx"))
        wb = openpyxl.load_workbook(init, read_only=True)
        dims = {ws.title: (ws.max_row, ws.max_column) for ws in wb.worksheets}
        wb.close()
        oversize = [f"{s}:{r}x{c}" for s, (r, c) in dims.items() if r > SER_ROWS or c > SER_COLS]

        print(f"\n{'=' * 78}")
        print(f"TASK {tid}  [{item['type']}]  status={item['status']}  "
              f"cells {item.get('correct', '?')}/{item.get('cells', '?')}")
        print(f"answer: {task.get('answer_sheet')}!{task['answer_position']}   data: {task.get('data_position')}")
        print(f"sheets: {dims}{'   OVERSIZE vs 120x30: ' + ', '.join(oversize) if oversize else ''}")
        print(f"instruction: {task['instruction'][:260]!r}")
        print(f"trace: out_tokens={trace['output_tokens']}"
              f"{' == CAP (truncated)' if trace['output_tokens'] in (8192, 24576) else ''}"
              f"  error={trace['error'][:90] if trace['error'] else None}")
        for m in item.get("mismatches", []):
            print(f"  {m['cell']}: expected {m['expected']!r} got {m['actual']!r}"
                  f"   <- {mismatch_flavor(m['expected'], m['actual'])}")
        if not trace["error"] and trace["response"]:
            print(f"reply tail: {trace['response'][-220:]!r}")


if __name__ == "__main__":
    main()
