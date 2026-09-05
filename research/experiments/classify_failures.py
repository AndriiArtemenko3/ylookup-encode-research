"""Automatic failure-pattern classification, and dev-vs-400 scale validation.

    uv run experiments/classify_failures.py ../experiments/E008-cascade-dev
    uv run experiments/classify_failures.py ../experiments/E008-cascade-dev ../experiments/E009-champion-400

Classifies every failed task of a run into the taxonomy first derived from
E002's 23 dev failures, using the same evidence rules, so distributions are
comparable across runs and scales:

  merged-cell     write crashed on a merged range (error text)
  truncated       a selected/only attempt consumed the whole token budget
  date-as-text    all mismatches have str(expected) == str(actual) (typed vs text)
  window          answer cells or referenced sheets exceed the 120x30 baseline view
  formula-error   Excel error values in produced answer cells
  wrong-values    everything else: completed, parsed, simply incorrect

With two runs it prints both distributions side by side (percentages of each
run's failures) — the scale check that dev-derived priorities hold on the 400.
Reporting tool only; reads official results.json + traces, never golden files.
"""

import json
import sys
from pathlib import Path

RESEARCH = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(RESEARCH))

import openpyxl
from openpyxl.utils.cell import coordinate_to_tuple

from sb import answer_cells, load_dataset

ERROR_VALUES = {"#NAME?", "#N/A", "#REF!", "#VALUE!", "#DIV/0!", "#NULL!", "#NUM!"}
ORDER = ["date-as-text", "window", "truncated", "formula-error", "wrong-values", "merged-cell", "unscored"]


def classify(item: dict, task: dict, traces: list[dict], max_tokens: int) -> str:
    if item.get("status") not in ("graded",):
        blob = json.dumps(traces)
        if "MergedCell" in blob:
            return "merged-cell"
        return "unscored"
    selected = next((t for t in traces if t.get("selected")), traces[-1] if traces else {})
    err_text = " ".join((t.get("error") or "") for t in traces)
    if "MergedCell" in err_text:
        return "merged-cell"
    mismatches = item.get("mismatches") or []
    if mismatches and all(str(m["expected"]) == str(m["actual"]) for m in mismatches):
        return "date-as-text"
    if any(isinstance(m.get("actual"), str) and m["actual"].strip() in ERROR_VALUES for m in mismatches):
        return "formula-error"
    if (selected.get("output_tokens") or 0) >= max_tokens:
        return "truncated"
    wb = openpyxl.load_workbook(task["init_xlsx"], read_only=True)
    cells = answer_cells(task, wb)
    oversize = any((ws.max_row or 0) > 120 or (ws.max_column or 0) > 30 for ws in wb.worksheets)
    wb.close()
    outside = any(coordinate_to_tuple(c)[0] > 120 or coordinate_to_tuple(c)[1] > 30 for _s, c in cells)
    if outside or (oversize and any(m.get("actual") in (None, 0, "0") for m in mismatches)):
        return "window"
    return "wrong-values"


def run_distribution(exp_dir: Path, tasks: dict[str, dict]) -> tuple[dict, int]:
    results = json.loads((exp_dir / "results.json").read_text())
    manifest = json.loads((exp_dir / "manifest.json").read_text()) if (exp_dir / "manifest.json").exists() else {}
    max_tokens = manifest.get("max_tokens") or 24576
    counts = dict.fromkeys(ORDER, 0)
    failed_ids = []
    for item in results["items"]:
        if item.get("pass"):
            continue
        tid = str(item["id"])
        trace_file = exp_dir / "traces" / f"{tid}.jsonl"
        traces = [json.loads(line) for line in trace_file.read_text().splitlines()] if trace_file.exists() else []
        counts[classify(item, tasks[tid], traces, max_tokens)] += 1
        failed_ids.append(tid)
    return counts, len(failed_ids)


def main():
    dirs = [Path(a) for a in sys.argv[1:]]
    if not dirs:
        sys.exit("usage: classify_failures.py <experiment-dir> [<experiment-dir-2>]")
    tasks = {str(t["id"]): t for t in load_dataset()}
    rows = [(d.name, *run_distribution(d, tasks)) for d in dirs]
    width = max(len(p) for p in ORDER) + 2
    header = f"{'pattern':<{width}}" + "".join(f"{name[:24]:>26}" for name, _c, _n in rows)
    print(header)
    print("-" * len(header))
    for pattern in ORDER:
        line = f"{pattern:<{width}}"
        for _name, counts, n_failed in rows:
            c = counts[pattern]
            pct = f"({c / n_failed:５.0%})".replace("５", "") if n_failed else ""
            line += f"{c:>18} {pct:>7}"
        print(line)
    print("-" * len(header))
    line = f"{'failed total':<{width}}"
    for _name, _counts, n_failed in rows:
        line += f"{n_failed:>18} {'':>7}"
    print(line)


if __name__ == "__main__":
    main()
