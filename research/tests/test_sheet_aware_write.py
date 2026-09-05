"""Worker A/D unit tests: multi-sheet addressing and patch merge.

Run: uv run tests/test_sheet_aware_write.py
"""

import sys
import tempfile
import unittest
from pathlib import Path

RESEARCH = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(RESEARCH))
sys.path.insert(0, str(RESEARCH / "baseline"))

import openpyxl
from common import SpreadsheetAnswer

from inference.patch import apply_patch
from inference.write import split_cell_address, write_output


def make_task(tmp: Path):
    wb = openpyxl.Workbook()
    ws1 = wb.active; ws1.title = "Alpha"
    ws2 = wb.create_sheet("Beta")
    for ws in (ws1, ws2):
        ws["A1"] = "hdr"; ws["B6"] = "old"
    init = tmp / "init.xlsx"; wb.save(init)
    return {"id": "t", "init_xlsx": str(init),
            "answer_position": "'Alpha'!B6,'Beta'!B6", "answer_sheet": None}


class SheetAwareWrite(unittest.TestCase):
    def test_split(self):
        self.assertEqual(split_cell_address("Sheet2!b6"), ("sheet2", "B6"))
        self.assertEqual(split_cell_address("'My Sheet'!C2"), ("my sheet", "C2"))
        self.assertEqual(split_cell_address("b6"), (None, "B6"))

    def test_same_coordinate_two_sheets_distinct_values(self):
        with tempfile.TemporaryDirectory() as td:
            tmp = Path(td); task = make_task(tmp)
            answer = SpreadsheetAnswer.model_validate({"cells": [
                {"cell": "Alpha!B6", "value": 111}, {"cell": "Beta!B6", "value": 222}]})
            out = tmp / "out.xlsx"; write_output(task, answer, out)
            wb = openpyxl.load_workbook(out)
            self.assertEqual(wb["Alpha"]["B6"].value, 111)
            self.assertEqual(wb["Beta"]["B6"].value, 222)

    def test_old_plain_format_still_works(self):
        with tempfile.TemporaryDirectory() as td:
            tmp = Path(td); task = make_task(tmp)
            task["answer_position"] = "'Alpha'!B6"
            answer = SpreadsheetAnswer.model_validate({"cells": [{"cell": "B6", "value": 7}]})
            out = tmp / "out.xlsx"; write_output(task, answer, out)
            wb = openpyxl.load_workbook(out)
            self.assertEqual(wb["Alpha"]["B6"].value, 7)
            self.assertEqual(wb["Beta"]["B6"].value, "old")  # untouched

    def test_plain_format_on_multisheet_writes_both(self):
        # legacy behaviour preserved when the model does not qualify
        with tempfile.TemporaryDirectory() as td:
            tmp = Path(td); task = make_task(tmp)
            answer = SpreadsheetAnswer.model_validate({"cells": [{"cell": "B6", "value": 9}]})
            out = tmp / "out.xlsx"; write_output(task, answer, out)
            wb = openpyxl.load_workbook(out)
            self.assertEqual(wb["Alpha"]["B6"].value, 9)
            self.assertEqual(wb["Beta"]["B6"].value, 9)


class PatchMerge(unittest.TestCase):
    def test_patch_only_allowed_cells(self):
        with tempfile.TemporaryDirectory() as td:
            tmp = Path(td); task = make_task(tmp)
            base = SpreadsheetAnswer.model_validate({"cells": [
                {"cell": "Alpha!B6", "value": 1}, {"cell": "Beta!B6", "value": 2}]})
            prev = tmp / "prev.xlsx"; write_output(task, base, prev)
            patch = SpreadsheetAnswer.model_validate({"cells": [
                {"cell": "Beta!B6", "value": 99},       # allowed
                {"cell": "Alpha!B6", "value": 1000},    # NOT allowed -> rejected
                {"cell": "Beta!A1", "value": "evil"}]}) # NOT allowed -> rejected
            out = tmp / "patched.xlsx"
            applied, rejected = apply_patch(prev, patch, {("beta", "B6")}, out)
            self.assertEqual((applied, rejected), (1, 2))
            wb = openpyxl.load_workbook(out)
            self.assertEqual(wb["Beta"]["B6"].value, 99)
            self.assertEqual(wb["Alpha"]["B6"].value, 1)
            self.assertEqual(wb["Beta"]["A1"].value, "hdr")


if __name__ == "__main__":
    sys.exit(unittest.main(verbosity=2))
