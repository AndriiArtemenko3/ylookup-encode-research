"""Golden-free health assessment of OUR OWN outputs.

Inputs are things the pipeline legitimately knows at inference time: the trace
of the attempt (token counts, parse mode), the workbook we wrote, and — when
formulas were written — a LibreOffice recalculation of that workbook. It never
reads expected answers; it detects DISTRESS, not wrongness:

- truncated: the reply consumed the entire token budget (cut off mid-answer)
- parse_failed: no usable JSON at all (output falls back to the init copy)
- salvaged: strict parse failed; entries were recovered from a damaged reply
- static formula issues: unknown function names, references to nonexistent
  sheets, unbalanced parentheses (caught before paying for a recalculation)
- recalc_errors: Excel error values (#NAME?, #N/A, ...) in the answer cells
  of our recalculated output

Selection policy: hard failures dominate; the champion attempt wins ties.
Emptiness is reported but never acted on — clearing cells is sometimes the
correct answer, so "wrote little" must not trigger escalation by itself.
"""

import re
import sys
import tempfile
from dataclasses import dataclass, field
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import openpyxl

from sb import answer_cells, load_answer_values, recalculate

ERROR_VALUES = {"#NAME?", "#N/A", "#REF!", "#VALUE!", "#DIV/0!", "#NULL!", "#NUM!"}

# Classic functions that need no prefix, for the static unknown-function check.
_CLASSIC_FUNCS = {
    "SUM", "AVERAGE", "MIN", "MAX", "COUNT", "COUNTA", "COUNTBLANK", "PRODUCT",
    "SUMIF", "SUMIFS", "COUNTIF", "COUNTIFS", "AVERAGEIF", "AVERAGEIFS", "SUMPRODUCT",
    "IF", "AND", "OR", "NOT", "IFERROR", "IFNA", "ISNUMBER", "ISTEXT", "ISBLANK", "ISERROR",
    "INDEX", "MATCH", "VLOOKUP", "HLOOKUP", "LOOKUP", "OFFSET", "INDIRECT", "CHOOSE",
    "LEFT", "RIGHT", "MID", "LEN", "TRIM", "UPPER", "LOWER", "PROPER", "SUBSTITUTE",
    "REPLACE", "SEARCH", "FIND", "TEXT", "VALUE", "CONCATENATE", "EXACT", "REPT", "CHAR", "CODE",
    "DATE", "YEAR", "MONTH", "DAY", "HOUR", "MINUTE", "SECOND", "TIME", "TODAY", "NOW",
    "EOMONTH", "EDATE", "WEEKDAY", "DATEDIF", "DATEVALUE", "TIMEVALUE", "NETWORKDAYS", "WORKDAY",
    "ROUND", "ROUNDUP", "ROUNDDOWN", "INT", "TRUNC", "ABS", "MOD", "CEILING", "FLOOR",
    "RANK", "LARGE", "SMALL", "MEDIAN", "MODE", "STDEV", "VAR", "PERCENTILE", "QUARTILE",
    "ROW", "ROWS", "COLUMN", "COLUMNS", "TRANSPOSE", "N", "T", "NA",
    "IRR", "XIRR", "NPV", "XNPV", "PMT", "PV", "FV", "RATE", "POWER", "SQRT", "EXP", "LN", "LOG",
}
_FUNC_CALL = re.compile(r"(?<![A-Z0-9_.])([A-Z][A-Z0-9]{1,15})\(")
_SHEET_REF = re.compile(r"'([^']+)'!|(?<![A-Z0-9_.:'])([A-Za-z_][A-Za-z0-9_.]{0,30})!")


@dataclass
class HealthReport:
    stage: str
    truncated: bool = False
    parse_failed: bool = False
    salvaged: bool = False
    static_issues: list = field(default_factory=list)
    recalc_errors: int = 0
    recalc_ran: bool = False
    answered: int = 0
    total: int = 0

    @property
    def hard_failures(self) -> int:
        """Weighted count of definite problems; 0 == healthy."""
        return (8 * self.parse_failed + 4 * self.truncated
                + 2 * min(self.recalc_errors, 4) + 2 * min(len(self.static_issues), 2)
                + 1 * self.salvaged)

    @property
    def healthy(self) -> bool:
        return self.hard_failures == 0

    def summary(self) -> dict:
        return {"stage": self.stage, "healthy": self.healthy, "hard_failures": self.hard_failures,
                "truncated": self.truncated, "parse_failed": self.parse_failed,
                "salvaged": self.salvaged, "static_issues": self.static_issues[:4],
                "recalc_errors": self.recalc_errors, "recalc_ran": self.recalc_ran,
                "answered": self.answered, "total": self.total}


def static_formula_issues(formulas: list[str], sheet_names: set[str],
                          known_prefixed: set[str]) -> list[str]:
    issues = []
    for f in formulas:
        if f.count("(") != f.count(")"):
            issues.append(f"unbalanced parens: {f[:60]}")
        for m in _FUNC_CALL.finditer(f.upper()):
            name = m.group(1)
            if name not in _CLASSIC_FUNCS and name not in known_prefixed and not f.upper().count(f"_XLFN.{name}"):
                issues.append(f"unknown function {name}: {f[:60]}")
                break
        for m in _SHEET_REF.finditer(f):
            sheet = (m.group(1) or m.group(2) or "").strip()
            if sheet and sheet not in sheet_names:
                issues.append(f"unknown sheet {sheet!r}: {f[:60]}")
                break
    return issues


def assess(task: dict, output_xlsx: Path, *, stage: str, output_tokens: int | None,
           max_tokens: int, parse_mode: str | None) -> HealthReport:
    """parse_mode: 'strict' | 'salvaged' | None (None = parse failed entirely)."""
    report = HealthReport(stage=stage)
    report.parse_failed = parse_mode is None
    report.salvaged = parse_mode == "salvaged"
    report.truncated = output_tokens is not None and output_tokens >= max_tokens

    try:
        wb = openpyxl.load_workbook(output_xlsx)
        sheet_names = set(wb.sheetnames)
        formulas = []
        for sheet, coord in answer_cells(task, wb):
            ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
            v = ws[coord].value
            report.total += 1
            if v is not None:
                report.answered += 1
            if isinstance(v, str) and v.startswith("="):
                formulas.append(v)
        wb.close()

        if formulas:
            from inference.write import _XLFN_FUNCS, _XLWS_FUNCS
            report.static_issues = static_formula_issues(formulas, sheet_names, _XLFN_FUNCS | _XLWS_FUNCS)
            if not report.static_issues:
                with tempfile.TemporaryDirectory() as work:
                    recalced = recalculate(output_xlsx, work)
                    values = load_answer_values(recalced, task)
                report.recalc_ran = True
                report.recalc_errors = sum(1 for v in values.values()
                                           if isinstance(v, str) and v.strip() in ERROR_VALUES)
    except Exception as e:
        report.static_issues.append(f"health check error: {type(e).__name__}: {e}"[:120])
    return report
