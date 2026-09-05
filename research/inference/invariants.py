"""H3: golden-free structural invariants over OUR OWN output workbook.

Detects defects a spreadsheet-literate human would flag without knowing the
answer: requested cells left empty while the rest of the range is filled, and
cells whose type conflicts with their column's majority type (text placeholder
in a numeric column, a date written as text among real datetimes). Emits
human-readable defect descriptions for a targeted repair prompt. Never reads
expected answers; rules are generic invariants developed on train tasks.
"""

import datetime
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import openpyxl

from sb import answer_cells

MAX_LISTED = 25
_DATEISH = re.compile(r"^\d{1,2}[/.-]\d{1,2}[/.-]\d{2,4}([ T]\d{1,2}:\d{2}(:\d{2})?)?$")


def _type_of(v):
    if v is None:
        return None
    if isinstance(v, bool):
        return "bool"
    if isinstance(v, (int, float)):
        return "number"
    if isinstance(v, (datetime.datetime, datetime.date, datetime.time)):
        return "datetime"
    return "text"


def structural_defects(task: dict, output_xlsx: Path) -> tuple[list[str], set]:
    """Returns (defect descriptions, flagged coords). Flagged coords + currently-
    empty cells are the ONLY cells a structural repair may change (acceptance
    guard in the cascade) — so a false-positive costs a wasted call, never a
    regression on untouched cells."""
    wb = openpyxl.load_workbook(output_xlsx, data_only=False)
    init = openpyxl.load_workbook(task["init_xlsx"], data_only=True)
    by_col = defaultdict(list)
    empties, filled, min_row = [], 0, {}
    for sheet, coord in answer_cells(task, wb):
        ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
        iws = init[ws.title] if ws.title in init.sheetnames else init.active
        v = ws[coord].value
        col = "".join(ch for ch in coord if ch.isalpha())
        row = int("".join(ch for ch in coord if ch.isdigit()))
        key = (ws.title, col)
        by_col[key].append((coord, v, iws[coord].value, row))
        min_row[key] = min(min_row.get(key, row), row)
        if v is None:
            # a skipped cell, not a cleared one: only rows that visibly hold data
            # elsewhere in the init sheet count as suspicious
            row_active = any(c.value is not None for c in iws[row])
            if row_active:
                empties.append(f"{ws.title}!{coord}")
        else:
            filled += 1
    init.close()
    wb.close()

    defects, flagged = [], set()
    total = filled + len(empties)
    # R1b: the answer wrote NOTHING at all — either wrong (skipped work) or the
    # rare clear-everything task; the repair prompt asks it to fill or confirm
    if filled == 0 and total >= 1:
        defects.append("your answer left EVERY cell in the answer range empty — produce the actual "
                       "values the instruction asks for, or return nulls only if the range must truly be empty")
        for sheet, coord in [(s, c) for (s, col) in by_col for c, _v, _iv, _r in by_col[(s, col)]][:200]:
            flagged.add((sheet, coord))
    # R1 completeness: empty cells on data-bearing rows while a solid share of the range is filled
    elif empties and total >= 4 and filled / total >= 0.3:
        listed = ", ".join(empties[:MAX_LISTED])
        more = f" (and {len(empties) - MAX_LISTED} more)" if len(empties) > MAX_LISTED else ""
        defects.append(f"these answer cells were left empty although their rows contain data: "
                       f"{listed}{more} — fill them, or return null only if they must truly be empty")
        for e in empties:
            sheet, coord = e.split("!", 1)
            flagged.add((sheet, coord))

    # R2 column-type consistency: minority text inside a typed column — but never
    # the range's first row (headers) and never cells whose init value was already
    # this text (echoing the input is always safe)
    for (sheet, col), cells in by_col.items():
        types = Counter(t for _c, v, _iv, _r in cells for t in [_type_of(v)] if t)
        if len(cells) < 4 or not types:
            continue
        majority, maj_n = types.most_common(1)[0]
        if maj_n / sum(types.values()) < 0.7 or majority not in ("number", "datetime"):
            continue
        for coord, v, init_v, row in cells:
            if (v is None or _type_of(v) != "text" or str(v).startswith("=")
                    or row == min_row[(sheet, col)] or (init_v is not None and str(init_v) == str(v))):
                continue
            if majority == "number":
                defects.append(f"{sheet}!{coord} contains text {str(v)[:30]!r} in a column of numbers — "
                               f"replace it with the correct numeric value (use 0 only if that is the value)")
                flagged.add((sheet, coord))
            elif _DATEISH.match(str(v).strip()):
                defects.append(f"{sheet}!{coord} contains the date-like TEXT {str(v)[:30]!r} in a column of "
                               f"real dates — rewrite it in ISO format (YYYY-MM-DD HH:MM:SS) so it becomes a true date")
                flagged.add((sheet, coord))
            if len(flagged) > 40:
                break
    return defects, flagged
