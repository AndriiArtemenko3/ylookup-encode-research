"""Harness v1 write path: write model answers to the workbook faithfully.

The model can only emit JSON strings/numbers; Excel cells are typed. The
official comparison (sb.transform_value / values_equal) requires equal types
after normalisation, so a date the model answered correctly as the string
"2014-02-05 00:00:00" fails against a real datetime cell. E002 lost 9 of 60
dev tasks to exactly this.

coerce_value converts ONLY strict, full-string ISO shapes — the forms our own
serialisation prints datetime/time cells in, i.e. what the model echoes back
when the answer is a date. Anything else (partial matches, display formats
like "2021 02 24", AM/PM strings) is left verbatim: when a golden wants date
LOOKING text, it wants the text. Regression scan on E002: zero passing answer
cells contained these shapes.

harness log: v1 = baseline + coerce_value in the write path. Nothing else.
"""

import datetime
import re
import shutil
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import openpyxl

from sb import answer_cells

HARNESS_VERSION = "h4-formulas"

# Modern Excel functions need the prefix Excel itself stores in the file, or
# both Excel and the LibreOffice recalculation return #NAME? (research/README.md:32).
_XLWS_FUNCS = {"FILTER", "SORT", "SORTBY"}
_XLFN_FUNCS = {
    "XLOOKUP", "XMATCH", "UNIQUE", "LET", "LAMBDA", "CHOOSECOLS", "CHOOSEROWS",
    "TEXTJOIN", "IFS", "MAXIFS", "MINIFS", "CONCAT", "SWITCH", "SEQUENCE",
    "RANDARRAY", "TEXTSPLIT", "TEXTBEFORE", "TEXTAFTER", "TOCOL", "TOROW",
    "VSTACK", "HSTACK", "TAKE", "DROP", "BYROW", "BYCOL", "MAP", "SCAN", "REDUCE",
}
_FUNC_CALL = re.compile(r"(?<![A-Z0-9_.])([A-Z][A-Z0-9]*)\(")


def fix_formula(formula: str) -> str:
    """Add the storage prefixes modern functions need to survive recalculation."""
    def repl(m):
        name = m.group(1)
        if name in _XLWS_FUNCS:
            return f"_xlfn._xlws.{name}("
        if name in _XLFN_FUNCS:
            return f"_xlfn.{name}("
        return m.group(0)

    return _FUNC_CALL.sub(repl, formula) if "_xlfn" not in formula else formula

_ISO_DATETIME = re.compile(r"^(\d{4})-(\d{2})-(\d{2})[ T](\d{2}):(\d{2}):(\d{2})$")
_ISO_DATE = re.compile(r"^(\d{4})-(\d{2})-(\d{2})$")
_ISO_TIME = re.compile(r"^(\d{1,2}):(\d{2}):(\d{2})$")


def coerce_value(value):
    """Strict full-match ISO datetime/date/time strings -> typed objects; all else verbatim."""
    if not isinstance(value, str):
        return value
    s = value.strip()
    if m := _ISO_DATETIME.match(s):
        try:
            return datetime.datetime(*map(int, m.groups()))
        except ValueError:  # e.g. month 13: not a date, keep the text
            return value
    if m := _ISO_DATE.match(s):
        try:
            return datetime.datetime(*map(int, m.groups()))
        except ValueError:
            return value
    if m := _ISO_TIME.match(s):
        h, mi, sec = map(int, m.groups())
        if h < 24 and mi < 60 and sec < 60:
            return datetime.time(h, mi, sec)
    return value


def _unmerge_target_ranges(wb, targets) -> None:
    """Unmerge only merged ranges that overlap cells we are about to write.

    The baseline crashes with "MergedCell.value is read-only" and loses the
    whole task; the failure mode here at worst matches that (a wrong workbook
    is no worse than the guaranteed-fail fallback).
    """
    from openpyxl.utils.cell import coordinate_to_tuple

    by_sheet: dict[str, tuple] = {}
    for ws, coord in targets:
        by_sheet.setdefault(ws.title, (ws, set()))[1].add(coordinate_to_tuple(coord))
    for ws, coords in by_sheet.values():
        for rng in list(ws.merged_cells.ranges):
            if any(rng.min_row <= r <= rng.max_row and rng.min_col <= c <= rng.max_col
                   for r, c in coords):
                ws.unmerge_cells(str(rng))


def split_cell_address(cell: str) -> tuple[str | None, str]:
    """'Sheet2!B6' -> ('sheet2', 'B6'); 'B6' -> (None, 'B6'). Sheet matched
    case-insensitively; quotes stripped. Fixes the multi-sheet coordinate
    collision: identical coordinates on different sheets carry distinct values."""
    if "!" in cell:
        sheet, coord = cell.rsplit("!", 1)
        return sheet.strip().strip("'\"").lower() or None, coord.strip().upper()
    return None, cell.strip().upper()


def write_output(task: dict, answer, out_path: Path) -> None:
    """Baseline write_output (baseline/common.py:79) plus coerce_value, unmerge,
    and sheet-aware answer addressing (coordinate-only entries stay valid)."""
    qualified, plain = {}, {}
    for c in answer.cells:
        sheet, coord = split_cell_address(c.cell)
        if sheet is not None:
            qualified[(sheet, coord)] = c.value
        else:
            plain[coord] = c.value
    shutil.copy(task["init_xlsx"], out_path)
    wb = openpyxl.load_workbook(out_path)
    targets = []
    for sheet, coord in answer_cells(task, wb):
        ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
        key = (ws.title.lower(), coord)
        if key in qualified:
            targets.append((ws, coord, qualified[key]))
        elif coord in plain:
            targets.append((ws, coord, plain[coord]))
    _unmerge_target_ranges(wb, [(ws, coord) for ws, coord, _v in targets])
    for ws, coord, value in targets:
        if isinstance(value, str) and value.startswith("="):
            ws[coord] = fix_formula(value)  # openpyxl stores "=..." as a formula
        else:
            ws[coord] = coerce_value(value)
    wb.save(out_path)
