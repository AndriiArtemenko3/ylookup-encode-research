"""Harness v3 serialisation: the model must be able to see the data it is asked about.

Fixes E002 taxonomy pattern 2 (7/60 dev tasks): the baseline window
(sb.serialize_workbook, 120 rows x 30 cols) silently hides rows/columns the
task needs — e.g. an answer range B1:B260 goes invisible at exactly row 121,
and a task needing column GK (col 193) sees columns A..AD only.

Rules:
1. The answer range's rows and columns are ALWAYS included, whatever else is.
2. The default window is much larger (500x60), sized against the model's 65k
   context; E002 used ~2.3k input tokens of it on average.
3. A sheet larger than the window gets a structured view — head rows, the
   answer block with context, tail rows — with EXPLICIT omission markers, so
   the model knows it is looking at a sample instead of mistaking it for the
   whole sheet.
4. A global character budget shrinks the window progressively for monster
   workbooks; answer rows/columns are never dropped.
"""

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries

from sb import answer_ranges

SERIALIZE_VERSION = "coverage-v1"
CHAR_BUDGET = 100_000  # ~25k tokens, leaves room for instruction + 24k output
WINDOW_STEPS = [(500, 60), (250, 40), (120, 30)]
ANSWER_CONTEXT_ROWS = 2
HEAD_ROWS = 30
TAIL_ROWS = 5


def _answer_rows_cols(task, sheet_title, active_title):
    rows, cols = set(), set()
    for sheet, rng in answer_ranges(task):
        target = sheet or active_title
        if target != sheet_title:
            continue
        min_col, min_row, max_col, max_row = range_boundaries(rng)
        rows.update(range(min_row or 1, (max_row or min_row or 1) + 1))
        cols.update(range(min_col, max_col + 1))
    return rows, cols


def _select(total, window, must_have, head, tail, context):
    """Indices to show: 1..window prefix, plus must-have blocks with context,
    plus head/tail anchors when the prefix does not already cover everything."""
    if total <= window:
        return list(range(1, total + 1))
    keep = set(range(1, min(head, window) + 1))
    keep.update(range(max(1, total - tail + 1), total + 1))
    budget = window - len(keep)
    prefix = [i for i in range(1, total + 1) if i not in keep]
    for m in sorted(must_have):
        for i in range(max(1, m - context), min(total, m + context) + 1):
            if i not in keep:
                keep.add(i)
                budget -= 1
    for i in prefix:
        if budget <= 0:
            break
        if i not in keep:
            keep.add(i)
            budget -= 1
    return sorted(i for i in keep if 1 <= i <= total)


MAX_FORMULA_CHARS = 90


def _formula_text(raw):
    """openpyxl gives '=...' strings, or ArrayFormula objects with .text."""
    raw = getattr(raw, "text", raw)
    if isinstance(raw, str) and raw.startswith("="):
        return raw if len(raw) <= MAX_FORMULA_CHARS else raw[:MAX_FORMULA_CHARS - 1] + "…"
    return None


def _serialize_sheet(ws, ws_formulas, task, active_title, max_rows, max_cols, formula_mode="all"):
    """formula_mode: 'all' annotates every formula cell; 'answer' only cells in
    the answer range's rows/columns (fill-downs of 1000s of formulas otherwise
    blow the char budget); 'none' matches the value-only view."""
    a_rows, a_cols = _answer_rows_cols(task, ws.title, active_title)
    rows = _select(ws.max_row, max_rows, a_rows, HEAD_ROWS, TAIL_ROWS, ANSWER_CONTEXT_ROWS)
    cols = _select(ws.max_column, max_cols, a_cols, max_cols, 0, 1)

    shown = f"showing {len(rows)}x{len(cols)} of {ws.max_row}x{ws.max_column}"
    lines = [f"### Sheet: {ws.title} ({shown})"]
    if len(cols) < ws.max_column:
        omitted = sorted(set(range(1, ws.max_column + 1)) - set(cols))
        lines.append(f"[columns omitted: {_ranges_note(omitted, get_column_letter)}]")
    header_at = len(lines)
    lines.append("\t".join([""] + [get_column_letter(c) for c in cols]))
    prev, any_formula = 0, False
    for r in rows:
        if r != prev + 1:
            lines.append(f"... rows {prev + 1}-{r - 1} omitted ({ws.max_row} total) ...")
        cells = []
        for c in cols:
            v = ws.cell(row=r, column=c).value
            text = "" if v is None else str(v)
            wanted = formula_mode == "all" or (formula_mode == "answer" and c in a_cols)
            formula = _formula_text(ws_formulas.cell(row=r, column=c).value) if (ws_formulas and wanted) else None
            if formula:
                text = f"{text} [{formula}]"
                any_formula = True
            cells.append(text)
        lines.append("\t".join([str(r)] + cells))
        prev = r
    if prev < ws.max_row:
        lines.append(f"... rows {prev + 1}-{ws.max_row} omitted ({ws.max_row} total) ...")
    if any_formula:
        lines.insert(header_at, "[cells holding formulas are shown as: cached_value [=formula]]")
    return "\n".join(lines)


def _ranges_note(indices, label):
    parts, start, prev = [], None, None
    for i in indices:
        if start is None:
            start = prev = i
        elif i == prev + 1:
            prev = i
        else:
            parts.append(label(start) if start == prev else f"{label(start)}-{label(prev)}")
            start = prev = i
    if start is not None:
        parts.append(label(start) if start == prev else f"{label(start)}-{label(prev)}")
    return ", ".join(parts)


def serialize_workbook(path, task):
    wb = openpyxl.load_workbook(path, data_only=True)
    try:  # second load with formulas intact; the init file is the task's own input
        wb_formulas = openpyxl.load_workbook(path, data_only=False)
    except Exception:
        wb_formulas = None
    steps = [(WINDOW_STEPS[0], "all")] + [(w, "answer") for w in WINDOW_STEPS] + [(WINDOW_STEPS[-1], "none")]
    for (max_rows, max_cols), formula_mode in steps:
        parts = []
        for ws in wb.worksheets:
            ws_f = wb_formulas[ws.title] if wb_formulas and ws.title in wb_formulas.sheetnames else None
            parts.append(_serialize_sheet(ws, ws_f, task, wb.active.title, max_rows, max_cols, formula_mode))
        text = "\n\n".join(parts)
        if len(text) <= CHAR_BUDGET:
            return text
    return text  # smallest window; answer rows/cols still guaranteed


def build_prompt(task: dict) -> str:
    """Baseline prompt shape (baseline/common.py:62) with the coverage serialiser."""
    return (
        f"## Instruction\n{task['instruction']}\n\n"
        f"## Workbook\n{serialize_workbook(task['init_xlsx'], task)}\n\n"
        f"## Answer range\nSheet: {task.get('answer_sheet') or 'active sheet'}\nCells: {task['answer_position']}\n"
    )
