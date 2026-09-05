"""Worker D: targeted patch repair — regenerate only suspicious cells.

26/95 E009 residual failures are within <=5 wrong cells; asking the model to
re-emit a 1000-cell answer to fix 3 cells is wasteful and risky. This module:

- build_patch_prompt: asks ONLY for the listed cells (sheet-qualified),
  given the instruction, the original workbook view, and per-cell evidence
  (structural defects, candidate disagreement proposals) — never goldens;
- apply_patch: merges a patch answer into a copy of the previous artifact,
  writing ONLY the requested cells (hard allowlist — a patch reply cannot
  touch anything else), through the standard coercion/formula write rules.

Integrates with H3 (defect coords), H9 (disagreement coords) and the cascade.
"""

import shutil
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import openpyxl

from inference.write import coerce_value, fix_formula, split_cell_address

PATCH_HINT = (
    "\n\nPATCH MODE: an earlier answer to this task is mostly correct. Re-derive ONLY these "
    "cells and return ONLY them, in the same JSON shape ({{\"cells\": [...]}}), using "
    "sheet-qualified addresses exactly as listed:\n{cells}\n"
    "Evidence per cell (from automatic validation and candidate disagreement — not from any "
    "answer key):\n{evidence}\n"
    "Do not include any other cells."
)


def build_patch_prompt(base_prompt: str, patch_cells: list[tuple[str, str]],
                       evidence: dict[tuple[str, str], str] | None = None) -> str:
    cells_txt = ", ".join(f"{s}!{c}" for s, c in patch_cells)
    ev = evidence or {}
    ev_txt = "\n".join(f"- {s}!{c}: {ev.get((s, c), 'flagged as suspicious')}"[:200]
                       for s, c in patch_cells[:25]) or "- (none)"
    return base_prompt + PATCH_HINT.format(cells=cells_txt, evidence=ev_txt)


def apply_patch(prev_artifact: Path, patch_answer, allowed: set[tuple[str, str]],
                out_path: Path) -> tuple[int, int]:
    """Copy prev artifact, write only patch cells inside `allowed`
    ((sheet_lower, COORD) pairs). Returns (applied, rejected)."""
    shutil.copy(prev_artifact, out_path)
    wb = openpyxl.load_workbook(out_path)
    sheets_lower = {name.lower(): name for name in wb.sheetnames}
    applied = rejected = 0
    for c in patch_answer.cells:
        sheet, coord = split_cell_address(c.cell)
        if sheet is None and len({s for s, _ in allowed}) == 1:
            sheet = next(iter({s for s, _ in allowed}))
        if sheet is None or (sheet, coord) not in allowed or sheet not in sheets_lower:
            rejected += 1
            continue
        ws = wb[sheets_lower[sheet]]
        value = c.value
        if isinstance(value, str) and value.startswith("="):
            ws[coord] = fix_formula(value)
        else:
            ws[coord] = coerce_value(value)
        applied += 1
    wb.save(out_path)
    return applied, rejected
