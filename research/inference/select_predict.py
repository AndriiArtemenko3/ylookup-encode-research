"""H4 risk-gated selector pipeline: h8 cascade + trajectory selection.

    uv run inference/select_predict.py --out-dir ... --base-model ... \
        [--ids ...] [--champion-cache DIR] [--group-size 3] [--concurrency 8]

Per task: run the h8 cascade; if the preregistered risk gate fires
(docs/H4_DEV_PLAN.md — unhealthy/escalated artifact, sheet-level type, or
answer range >= 10 cells), sample `group_size` extra champion-prompt
candidates at temperature 0.7 in ONE num_samples call, write each through the
h2 path, and select among {cascade artifact} ∪ {candidates} by the frozen
golden-free feature order (strict, no-error, closed, pairwise agreement,
answered). Ties go to the cascade artifact. Traces record every stage and the
selection decision.
"""

import argparse
import asyncio
import json
import os
import shutil
import sys
import time
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "baseline"))

import openpyxl
import tinker
from common import FORMAT_HINT, SYSTEM_PROMPT, append_jsonl, load_env, log, parse_ids, prepare_out_dir, selected_tasks
from common import build_prompt as champion_build_prompt
from tinker import types
from tinker_cookbook import renderers
from tinker_cookbook.model_info import get_recommended_renderer_name
from tinker_cookbook.tokenizer_utils import get_tokenizer

from inference.cascade import run_cascade_task
from inference.consensus import agreement_stats
from inference.parse import parse_answer_lenient
from inference.retry import with_backoff
from inference.write import write_output
from sb import DEFAULT_DATASET, answer_cells
from sb import answer_ranges as sb_answer_ranges

HARNESS_VERSION = "h9-selector"
GATE_MIN_CELLS = 10


def parse_args():
    p = argparse.ArgumentParser()
    p.add_argument("--out-dir", required=True)
    p.add_argument("--dataset-dir", default=str(DEFAULT_DATASET))
    p.add_argument("--ids")
    p.add_argument("--base-model", required=True)
    p.add_argument("--model-path")
    p.add_argument("--concurrency", type=int, default=8)
    p.add_argument("--max-tokens", type=int, default=24576)
    p.add_argument("--escalation-max-tokens", type=int, default=32768)
    p.add_argument("--group-size", type=int, default=3)
    p.add_argument("--champion-cache")
    return p.parse_args()


def parsed_cells_or_none(text):
    try:
        answer, _ = parse_answer_lenient(text)
        return {c.cell.upper(): c.value for c in answer.cells}, answer
    except Exception:
        return None, None


def artifact_cells(task, path):
    wb = openpyxl.load_workbook(path)
    out = {}
    for sheet, coord in answer_cells(task, wb):
        ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
        out[f"{ws.title}!{coord}".upper()] = ws[coord].value
    wb.close()
    return out


def gate_fires(task, cascade_traces):
    sel = next((t for t in cascade_traces if t.get("selected")), {})
    health = sel.get("health") or {}
    if not health.get("healthy", False):
        return "artifact unhealthy"
    if any(t.get("stage") not in ("champion",) for t in cascade_traces):
        return "escalation ran"
    if task["instruction_type"].startswith("Sheet"):
        return "sheet-level"
    wb = openpyxl.load_workbook(task["init_xlsx"], read_only=True)
    n = len(answer_cells(task, wb)); wb.close()
    if n >= GATE_MIN_CELLS:
        return f"answer cells {n} >= {GATE_MIN_CELLS}"
    return None


async def main():
    load_env()
    args = parse_args()
    service = tinker.ServiceClient(project_id=os.environ.get("TINKER_PROJECT_ID") or None)
    sampler = service.create_sampling_client(base_model=args.base_model, model_path=args.model_path)
    renderer = renderers.get_renderer(get_recommended_renderer_name(args.base_model), get_tokenizer(args.base_model))
    stops = renderer.get_stop_sequences()
    model = args.model_path or args.base_model

    async def complete(system, user_prompt, max_tokens):
        params = types.SamplingParams(max_tokens=max_tokens, temperature=0, stop=stops)
        messages = [{"role": "system", "content": system}, {"role": "user", "content": user_prompt + FORMAT_HINT}]
        mi = renderer.build_generation_prompt(messages)
        resp = await with_backoff(lambda: sampler.sample_async(prompt=mi, num_samples=1, sampling_params=params))
        toks = resp.sequences[0].tokens
        content = renderer.parse_response(toks)[0]["content"]
        if not isinstance(content, str):
            content = "".join(p.get("text", "") for p in content if p.get("type") == "text")
        return content, mi.length, len(toks)

    async def sample_trio(task, n):
        user = champion_build_prompt(task)
        sheets = {s for s, _r in sb_answer_ranges(task)}
        if len(sheets) > 1:
            user += ("\n\nIMPORTANT: the answer range spans multiple sheets. Write every cell "
                     "address sheet-qualified as 'SheetName!CELL'.")
        params = types.SamplingParams(max_tokens=args.max_tokens, temperature=0.7, stop=stops)
        messages = [{"role": "system", "content": SYSTEM_PROMPT}, {"role": "user", "content": user + FORMAT_HINT}]
        mi = renderer.build_generation_prompt(messages)
        resp = await with_backoff(lambda: sampler.sample_async(prompt=mi, num_samples=n, sampling_params=params))
        out = []
        for seq in resp.sequences:
            content = renderer.parse_response(seq.tokens)[0]["content"]
            if not isinstance(content, str):
                content = "".join(p.get("text", "") for p in content if p.get("type") == "text")
            out.append((content, len(seq.tokens)))
        return user, out

    tasks = selected_tasks(Path(args.dataset_dir), parse_ids(args.ids))
    out_dir = Path(args.out_dir)
    prepare_out_dir(out_dir)
    work = out_dir / "attempts"
    cache = Path(args.champion_cache) if args.champion_cache else None
    log(out_dir, f"harness {HARNESS_VERSION}  model {model}  tasks {len(tasks)}  N={args.group_size}")
    semaphore = asyncio.Semaphore(args.concurrency)

    async def run_one(task):
        async with semaphore:
            traces, status = [], "error: unknown"
            final = out_dir / "outputs" / f"{task['id']}.xlsx"
            try:
                best_path, traces, status = await run_cascade_task(
                    complete, task, work, max_tokens=args.max_tokens, model=model,
                    cache_dir=cache, escalation_max_tokens=args.escalation_max_tokens)
                shutil.copy(best_path, final)
                reason = gate_fires(task, traces)
                if reason:
                    started = time.time()
                    user, samples = await sample_trio(task, args.group_size)
                    tmp = work / task["id"]
                    pool = [{"name": "cascade", "path": final,
                             "cells": artifact_cells(task, final),
                             "strict": True,
                             "closed": True, "err": status.startswith("error")}]
                    for k, (content, ntok) in enumerate(samples):
                        cells, answer = parsed_cells_or_none(content)
                        rec = {"step": len(traces) + 1 + k, "model": model, "stage": f"h4-candidate-{k}",
                               "prompt": user, "response": content, "input_tokens": None,
                               "output_tokens": ntok, "latency_ms": int((time.time() - started) * 1000),
                               "error": None if cells else "parse failed"}
                        traces.append(rec)
                        if cells is None:
                            continue
                        cpath = tmp / f"h4cand{k}.xlsx"
                        await asyncio.to_thread(write_output, task, answer, cpath)
                        pool.append({"name": f"cand{k}", "path": cpath,
                                     "cells": artifact_cells(task, cpath),
                                     "strict": "{" in content and content.rstrip().endswith("}"),
                                     "closed": (ntok < args.max_tokens), "err": False})
                    stats = agreement_stats([p["cells"] for p in pool])
                    dis = stats["pairwise_disagreement"]
                    def score(i, p):
                        agr = 1 - sum(dis[i][j] for j in range(len(pool)) if j != i) / max(len(pool) - 1, 1)
                        return (p["strict"], not p["err"], p["closed"], round(agr, 4),
                                sum(v is not None for v in p["cells"].values()), -i)
                    best_i = max(range(len(pool)), key=lambda i: score(i, pool[i]))
                    chosen = pool[best_i]
                    for t in traces:
                        t["h4_gate"] = reason
                    traces[-1]["h4_selection"] = {"pool": len(pool), "chosen": chosen["name"],
                                                  "adopted": chosen["name"] != "cascade"}
                    if chosen["name"] != "cascade":
                        shutil.copy(chosen["path"], final)
                        status = "ok (h4-selected)"
            except Exception as e:
                if not final.exists():
                    shutil.copy(task["init_xlsx"], final)
                status = f"error: {e}"[:200]
                traces.append({"step": len(traces) + 1, "model": model, "stage": "h4-crash",
                               "prompt": None, "response": None, "input_tokens": None,
                               "output_tokens": None, "latency_ms": None,
                               "error": f"{type(e).__name__}: {e}"[:400]})
            for r in traces:
                append_jsonl(out_dir / "traces" / f"{task['id']}.jsonl", r)
            append_jsonl(out_dir / "predictions.jsonl",
                         {"id": task["id"], "output": f"outputs/{task['id']}.xlsx", "status": status})
        log(out_dir, f"{task['id']:<8} {status[:50]}")

    await asyncio.gather(*(run_one(t) for t in tasks))
    shutil.rmtree(work, ignore_errors=True)


if __name__ == "__main__":
    asyncio.run(main())
