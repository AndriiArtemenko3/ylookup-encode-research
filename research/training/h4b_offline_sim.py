"""H4B TRAIN mechanism test: executable-evidence selection vs H4 (offline).

    uv run training/h4b_offline_sim.py [--sample 140]

Implements exactly the FROZEN ordering of docs/H4B_VERIFIER_PLAN.md over the
existing F1-rollouts candidate pools. Goldens are used ONLY to score the
simulation afterwards (train-side analysis); never inside the selector.
"""

import argparse
import json
import random
import sys
import tempfile
from pathlib import Path

RESEARCH = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(RESEARCH))
sys.path.insert(0, str(RESEARCH / "baseline"))

import openpyxl

from inference.consensus import agreement_stats
from inference.health import ERROR_VALUES
from inference.parse import parse_answer_lenient
from inference.write import write_output
from sb import answer_cells, load_dataset, recalculate
from training.reward import train_ids

ROLLOUTS = RESEARCH.parent / "experiments" / "F1-rollouts" / "rollouts"


def expected_coords(task):
    wb = openpyxl.load_workbook(task["init_xlsx"], read_only=True)
    cells = answer_cells(task, wb)
    wb.close()
    return {c for _s, c in cells}


def candidate_features(task, record, work, expected):
    """FROZEN H4B evidence per the plan. Returns None if unparseable."""
    response = record.get("response") or ""
    try:
        answer, mode = parse_answer_lenient(response)
    except Exception:
        return None
    cells = {}
    for c in answer.cells:
        coord = c.cell.upper().rsplit("!", 1)[-1]
        cells[coord] = c.value
    path = Path(work) / f"cand{record['k']}.xlsx"
    try:
        write_output(task, answer, path)
        write_ok = True
    except Exception:
        return {"write_ok": False, "exec_ok": False, "coverage": 0.0, "stable": False,
                "strict": False, "closed": False, "cells": {}, "answered": 0}
    has_formula = any(isinstance(v, str) and str(v).startswith("=") for v in cells.values())
    target_path = path
    if has_formula:
        try:
            target_path = recalculate(path, work)
        except Exception:
            return {"write_ok": True, "exec_ok": False, "coverage": 0.0, "stable": False,
                    "strict": mode == "strict", "closed": True, "cells": cells, "answered": 0}
    wb = openpyxl.load_workbook(target_path, data_only=True)
    errors = 0
    values = {}
    for sheet, coord in answer_cells(task, wb):
        ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
        v = ws[coord].value
        values[coord] = v
        if isinstance(v, str) and v.strip() in ERROR_VALUES:
            errors += 1
    wb.close()
    represented = sum(1 for c in expected if c in cells)  # explicit nulls count
    coverage = represented / len(expected) if expected else 1.0
    cov_bucket = 2 if coverage >= 0.999 else 1 if coverage >= 0.9 else 0
    closed = response.rstrip().endswith("}")
    return {"write_ok": write_ok, "exec_ok": write_ok and errors == 0,
            "coverage": cov_bucket, "stable": True,  # literal round-trip stable by construction; formulas recalced once
            "strict": mode == "strict", "closed": closed,
            "cells": {("", k): v for k, v in values.items()},
            "answered": sum(v is not None for v in values.values())}


def h4_score(i, feats, dis):
    f = feats[i]
    agr = 1 - sum(dis[i][j] for j in range(len(feats)) if j != i) / max(len(feats) - 1, 1)
    return (f["strict"], f["write_ok"], f["closed"], round(agr, 4), f["answered"], -i)


def h4b_score(i, feats, dis):
    f = feats[i]
    agr = 1 - sum(dis[i][j] for j in range(len(feats)) if j != i) / max(len(feats) - 1, 1)
    return (f["exec_ok"], f["coverage"], f["stable"], f["strict"], f["closed"],
            round(agr, 4), f["answered"], -i)


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--sample", type=int, default=0)
    args = p.parse_args()
    tasks = {t["id"]: t for t in load_dataset() if t["id"] in train_ids()}

    task_dirs = sorted(d for d in ROLLOUTS.iterdir() if d.name in tasks)
    if args.sample:
        task_dirs = random.Random(7).sample(task_dirs, min(args.sample, len(task_dirs)))

    n = p1 = h4 = h4b = oracle = 0
    adoptions = adoption_wins = adoption_losses = 0
    fixed_by_exec = []
    for td in task_dirs:
        task = tasks[td.name]
        records = [json.loads(f.read_text()) for f in sorted(td.glob("[0-9]*.json"))]
        passes = [bool((r.get("score") or {}).get("pass")) for r in records]
        expected = expected_coords(task)
        with tempfile.TemporaryDirectory() as work:
            feats, keep = [], []
            for r in records:
                f = candidate_features(task, r, work, expected)
                if f is not None:
                    feats.append(f)
                    keep.append(r)
            if len(feats) < 3:
                continue
            n += 1
            kp = [bool((r.get("score") or {}).get("pass")) for r in keep]
            p1 += sum(kp) / len(kp)
            oracle += any(kp)
            stats = agreement_stats([f["cells"] for f in feats])
            dis = stats["pairwise_disagreement"]
            i4 = max(range(len(feats)), key=lambda i: h4_score(i, feats, dis))
            i4b = max(range(len(feats)), key=lambda i: h4b_score(i, feats, dis))
            h4 += kp[i4]
            h4b += kp[i4b]
            if i4b != i4:
                adoptions += 1
                if kp[i4b] and not kp[i4]:
                    adoption_wins += 1
                    fixed_by_exec.append(td.name)
                elif kp[i4] and not kp[i4b]:
                    adoption_losses += 1

    print(f"tasks simulated: {n}")
    print(f"pass@1:        {p1 / n:.3f}")
    print(f"H4 selected:   {h4 / n:.3f}")
    print(f"H4B selected:  {h4b / n:.3f}")
    print(f"oracle@pool:   {oracle / n:.3f}")
    gap = oracle - h4
    print(f"H4B recovers {((h4b - h4) / gap if gap else 0):.0%} of remaining H4->oracle headroom")
    print(f"divergent picks: {adoptions} | exec-fixed: {adoption_wins} {fixed_by_exec[:8]} | exec-broken: {adoption_losses}")


if __name__ == "__main__":
    main()
